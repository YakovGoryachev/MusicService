from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Avg
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.utils import timezone
from .models import (
    Track, Album, Playlist, Genre, TrackRating, AlbumRating, Comment, 
    User, Group, Artist, ArtistGroup, TrackGenre, PlaylistTrack
)
from .forms import UserRegistrationForm, UserLoginForm, PlaylistForm, CommentForm, TrackCreateForm
import json
from django.core.mail import send_mass_mail, EmailMessage
from django.http import HttpResponse
import io
from django.db.models import Count, Sum


def home(request):
    """Главная страница"""
    latest_tracks = Track.objects.select_related('album', 'album__artist', 'album__group').order_by('-id')[:8]
    popular_albums = Album.objects.select_related('artist', 'group').order_by('-play_count')[:4]
    genres = Genre.objects.all()[:6]
    
    context = {
        'latest_tracks': latest_tracks,
        'popular_albums': popular_albums,
        'genres': genres,
    }
    return render(request, 'music/home.html', context)


def track_list(request):
    """Список всех треков"""
    query = request.GET.get('q', '')
    genre_filter = request.GET.get('genre', '')
    artist_filter = request.GET.get('artist', '')
    group_filter = request.GET.get('group', '')
    
    tracks = Track.objects.select_related('album', 'album__artist', 'album__group').prefetch_related('genres', 'ratings')
    
    if query:
        tracks = tracks.filter(
            Q(name__icontains=query) |
            Q(album__name__icontains=query) |
            Q(album__artist__name__icontains=query) |
            Q(album__group__name__icontains=query) |
            Q(genres__name__icontains=query)
        ).distinct()
    
    if genre_filter:
        tracks = tracks.filter(genres__name=genre_filter)
    
    if artist_filter:
        tracks = tracks.filter(album__artist__name__icontains=artist_filter)
    
    if group_filter:
        tracks = tracks.filter(album__group__name__icontains=group_filter)
    
    tracks = tracks.order_by('-id')
    
    paginator = Paginator(tracks, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    genres = Genre.objects.all()
    
    context = {
        'tracks': page_obj,
        'genres': genres,
        'query': query,
        'genre_filter': genre_filter,
    }
    return render(request, 'music/track_list.html', context)


def track_detail(request, pk):
    """Детальная страница трека"""
    track = get_object_or_404(Track.objects.select_related('album', 'album__artist', 'album__group').prefetch_related('genres'), pk=pk)
    
    # Получаем оценки и комментарии
    ratings = TrackRating.objects.filter(track=track).select_related('user')
    comments = Comment.objects.filter(track=track).select_related('user').order_by('-created_at')
    
    # Средняя оценка
    avg_rating = ratings.aggregate(avg=Avg('value'))['avg'] or 0
    
    # Пользовательская оценка
    user_rating = None
    if request.user.is_authenticated:
        user_rating = ratings.filter(user=request.user).first()
    
    # Увеличиваем счетчик прослушиваний
    track.play_count += 1
    track.save()
    
    context = {
        'track': track,
        'ratings': ratings,
        'comments': comments,
        'avg_rating': round(avg_rating, 1),
        'user_rating': user_rating,
        'comment_form': CommentForm(),
    }
    return render(request, 'music/track_detail.html', context)


def album_list(request):
    """Список альбомов"""
    query = request.GET.get('q', '')
    
    albums = Album.objects.select_related('artist', 'group').prefetch_related('ratings')
    
    if query:
        albums = albums.filter(
            Q(name__icontains=query) |
            Q(artist__name__icontains=query) |
            Q(group__name__icontains=query)
        ).distinct()
    
    albums = albums.order_by('-release_date')
    
    paginator = Paginator(albums, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'albums': page_obj,
        'query': query,
    }
    return render(request, 'music/album_list.html', context)


def album_detail(request, pk):
    """Детальная страница альбома"""
    album = get_object_or_404(Album.objects.select_related('artist', 'group').prefetch_related('ratings'), pk=pk)
    tracks = album.track_set.all().prefetch_related('ratings', 'genres').order_by('id')
    
    # Получаем оценки альбома
    ratings = AlbumRating.objects.filter(album=album).select_related('user')
    
    # Средняя оценка альбома
    avg_rating = ratings.aggregate(avg=Avg('value'))['avg'] or 0
    
    # Пользовательская оценка альбома
    user_rating = None
    if request.user.is_authenticated:
        user_rating = ratings.filter(user=request.user).first()
    
    # Общая длительность альбома
    total_duration = sum(track.duration or 0 for track in tracks)
    
    context = {
        'album': album,
        'tracks': tracks,
        'ratings': ratings,
        'avg_rating': round(avg_rating, 1),
        'user_rating': user_rating,
        'total_duration': total_duration,
    }
    return render(request, 'music/album_detail.html', context)


def artist_list(request):
    """Список артистов"""
    query = request.GET.get('q', '')
    
    artists = (
        Artist.objects.all()
        .annotate(
            album_count=Count('album', distinct=True),
            track_count=Count('album__track', distinct=True),
        )
    )
    
    if query:
        artists = artists.filter(name__icontains=query)
    
    artists = artists.order_by('name')
    
    paginator = Paginator(artists, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'artists': page_obj,
        'query': query,
    }
    return render(request, 'music/artist_list.html', context)


def artist_detail(request, pk):
    """Детальная страница артиста"""
    artist = get_object_or_404(Artist, pk=pk)
    albums = artist.album_set.all().order_by('-release_date')
    tracks = Track.objects.filter(album__artist=artist).select_related('album').prefetch_related('genres').order_by('-play_count', '-id')
    artist_album_count = albums.count()
    artist_track_count = tracks.count()
    
    context = {
        'artist': artist,
        'albums': albums,
        'tracks': tracks,
        'artist_album_count': artist_album_count,
        'artist_track_count': artist_track_count,
    }
    return render(request, 'music/artist_detail.html', context)


def group_list(request):
    """Список групп"""
    query = request.GET.get('q', '')
    
    groups = Group.objects.all()
    
    if query:
        groups = groups.filter(name__icontains=query)
    
    groups = groups.order_by('name')
    
    paginator = Paginator(groups, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'groups': page_obj,
        'query': query,
    }
    return render(request, 'music/group_list.html', context)


def group_detail(request, pk):
    """Детальная страница группы"""
    group = get_object_or_404(Group, pk=pk)
    albums = group.album_set.all().order_by('-release_date')
    # Связь артиста и группы через ArtistGroup
    artists = Artist.objects.filter(artistgroup__group=group).distinct()
    artist_links = ArtistGroup.objects.filter(group=group).select_related('artist')
    # Треки группы через альбомы этой группы
    tracks_qs = Track.objects.filter(album__group=group).select_related('album').prefetch_related('genres', 'ratings').order_by('-play_count', '-id')
    tracks_count = tracks_qs.count()
    from django.db.models import Sum as _Sum
    total_play_count = tracks_qs.aggregate(total=_Sum('play_count'))['total'] or 0
    
    context = {
        'group': group,
        'albums': albums,
        'artists': artists,
        'artist_links': artist_links,
        'tracks': tracks_qs,
        'tracks_count': tracks_count,
        'total_play_count': total_play_count,
    }
    return render(request, 'music/group_detail.html', context)


def playlist_list(request):
    """Список публичных плейлистов"""
    query = request.GET.get('q', '')
    
    playlists = Playlist.objects.filter(is_public=True).select_related('user').prefetch_related('tracks')
    
    if query:
        playlists = playlists.filter(
            Q(name__icontains=query) |
            Q(description__icontains=query) |
            Q(user__login__icontains=query)
        ).distinct()
    
    paginator = Paginator(playlists, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'playlists': page_obj,
        'query': query,
    }
    return render(request, 'music/playlist_list.html', context)


def playlist_detail(request, pk):
    """Детальная страница плейлиста"""
    playlist = get_object_or_404(
        Playlist.objects.select_related('user').prefetch_related('tracks', 'playlist_tracks__track__album__artist', 'playlist_tracks__track__album__group', 'playlist_tracks__track__genres'), 
        pk=pk
    )
    
    if not playlist.is_public and request.user != playlist.user:
        messages.error(request, 'Этот плейлист приватный')
        return redirect('music:playlist_list')
    
    # Получаем треки через промежуточную таблицу для правильного порядка
    tracks = [pt.track for pt in playlist.playlist_tracks.all().order_by('added_date')]
    
    context = {
        'playlist': playlist,
        'tracks': tracks,
    }
    return render(request, 'music/playlist_detail.html', context)


@login_required
def my_playlists(request):
    """Мои плейлисты"""
    playlists = Playlist.objects.filter(user=request.user).prefetch_related('tracks')
    
    context = {
        'playlists': playlists,
    }
    return render(request, 'music/my_playlists.html', context)


@login_required
def create_playlist(request):
    """Создание плейлиста"""
    if not request.user.is_authenticated:
        messages.error(request, 'Необходимо войти в систему для создания плейлиста.')
        return redirect('music:login')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        is_private = request.POST.get('is_private') == 'True'
        photo = request.FILES.get('photo')
        
        if name:
            try:
                playlist = Playlist.objects.create(
                    user=request.user,
                    name=name,
                    description=description,
                    is_public=not is_private,
                    photo=photo,
                )
                messages.success(request, f'Плейлист "{name}" создан успешно!')
                return redirect('music:playlist_detail', pk=playlist.pk)
            except Exception as e:
                messages.error(request, f'Ошибка создания плейлиста: {str(e)}')
        else:
            messages.error(request, 'Название плейлиста обязательно!')
    
    # Получаем жанры для формы
    genres = Genre.objects.all()
    
    context = {
        'genres': genres,
    }
    return render(request, 'music/create_playlist.html', context)


@login_required
def edit_playlist(request, pk):
    """Редактирование плейлиста"""
    playlist = get_object_or_404(Playlist, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = PlaylistForm(request.POST, request.FILES, instance=playlist)
        if form.is_valid():
            playlist = form.save()
            # genres уже сохраняются ModelForm'ой через m2m, фото берётся из FILES
            messages.success(request, 'Плейлист обновлен!')
            return redirect('music:playlist_detail', pk=playlist.pk)
    else:
        form = PlaylistForm(instance=playlist)
    
    context = {
        'form': form,
        'playlist': playlist,
    }
    return render(request, 'music/edit_playlist.html', context)


@login_required
def delete_playlist(request, pk):
    """Удаление плейлиста"""
    playlist = get_object_or_404(Playlist, pk=pk, user=request.user)
    
    if request.method == 'POST':
        playlist.delete()
        messages.success(request, 'Плейлист удален!')
        return redirect('music:my_playlists')
    
    context = {
        'playlist': playlist,
    }
    return render(request, 'music/delete_playlist.html', context)


@login_required
@require_POST
def add_to_playlist(request, track_id):
    """Добавление трека в плейлист"""
    track = get_object_or_404(Track, pk=track_id)
    playlist_id = request.POST.get('playlist_id')
    
    if playlist_id:
        playlist = get_object_or_404(Playlist, pk=playlist_id, user=request.user)
        playlist.tracks.add(track)
        messages.success(request, f'Трек "{track.name}" добавлен в плейлист "{playlist.name}"!')
    else:
        messages.error(request, 'Выберите плейлист!')
            
    return redirect('music:track_detail', pk=track_id)


@login_required
@require_POST
def remove_from_playlist(request, playlist_id, track_id):
    """Удаление трека из плейлиста"""
    playlist = get_object_or_404(Playlist, pk=playlist_id, user=request.user)
    track = get_object_or_404(Track, pk=track_id)
    
    playlist.tracks.remove(track)
    messages.success(request, f'Трек "{track.name}" удален из плейлиста!')
    
    return redirect('music:playlist_detail', pk=playlist_id)


@login_required
@require_POST
def rate_track(request, track_id):
    """Оценка трека"""
    track = get_object_or_404(Track, pk=track_id)
    rating_value = request.POST.get('rating')
    
    if rating_value and rating_value.isdigit():
        rating_value = int(rating_value)
        if 1 <= rating_value <= 5:
            rating, created = TrackRating.objects.get_or_create(
                user=request.user,
                track=track,
                defaults={'value': rating_value}
            )
            if not created:
                rating.value = rating_value
                rating.save()
            
            messages.success(request, 'Оценка сохранена!')
        else:
            messages.error(request, 'Оценка должна быть от 1 до 5!')
    else:
        messages.error(request, 'Выберите оценку!')
            
    return redirect('music:track_detail', pk=track_id)


@login_required
@require_POST
def add_comment(request, track_id):
    """Добавление комментария к треку"""
    track = get_object_or_404(Track, pk=track_id)
    form = CommentForm(request.POST)
    
    if form.is_valid():
        comment = form.save(commit=False)
        comment.user = request.user
        comment.track = track
        comment.save()
        messages.success(request, 'Комментарий добавлен!')
    else:
        messages.error(request, 'Ошибка в форме комментария!')
    
    return redirect('music:track_detail', pk=track_id)


@login_required
def delete_comment(request, comment_id):
    """Удаление комментария"""
    comment = get_object_or_404(Comment, pk=comment_id, user=request.user)
    track_id = comment.track.id
    
    comment.delete()
    messages.success(request, 'Комментарий удален!')
    
    return redirect('music:track_detail', pk=track_id)


def user_register(request):
    """Регистрация пользователя"""
    if request.user.is_authenticated:
        return redirect('music:home')
    
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Регистрация прошла успешно!')
            return redirect('music:home')
    else:
        form = UserRegistrationForm()
    
    context = {
        'form': form,
    }
    return render(request, 'music/register.html', context)


def user_login(request):
    """Вход пользователя"""
    if request.user.is_authenticated:
        return redirect('music:home')
    
    if request.method == 'POST':
        form = UserLoginForm(request.POST)
        if form.is_valid():
            user_login = form.cleaned_data['login']
            password = form.cleaned_data['password']
            user = authenticate(request, username=user_login, password=password)
            
            if user is not None:
                if user.is_active:
                    login(request, user)
                    messages.success(request, f'Добро пожаловать, {user.login}!')
                    # Redirect to next URL if it exists, otherwise to home
                    next_url = request.GET.get('next')
                    return redirect(next_url if next_url else 'music:home')
                else:
                    form.add_error(None, 'Ваш аккаунт заблокирован. Пожалуйста, свяжитесь с администратором.')
            else:
                form.add_error(None, 'Неверный логин или пароль. Пожалуйста, проверьте введенные данные.')
    else:
        form = UserLoginForm()
    
    context = {
        'form': form,
        'next': request.GET.get('next', '')
    }
    return render(request, 'music/login.html', context)


def user_logout(request):
    """Выход пользователя"""
    logout(request)
    messages.success(request, 'Вы успешно вышли из системы!')
    return redirect('music:home')


@login_required
def profile(request):
    """Профиль пользователя"""
    from django.db.models import Count, Sum

    # Плейлисты пользователя
    user_playlists = Playlist.objects.filter(user=request.user).prefetch_related('tracks')
    recent_playlists = user_playlists.order_by('-creation_date')[:6]

    # Статистика
    playlists_count = user_playlists.count()
    total_tracks = sum(playlist.tracks.count() for playlist in user_playlists)
    total_duration = sum(
        sum(track.duration or 0 for track in playlist.tracks.all())
        for playlist in user_playlists
    )
    total_play_count = sum(
        sum(track.play_count for track in playlist.tracks.all())
        for playlist in user_playlists
    )

    # Оценки пользователя
    user_ratings = TrackRating.objects.filter(user=request.user).select_related('track', 'track__album')

    # Недавняя активность (оценки)
    recent_activity = []
    for rating in user_ratings.order_by('-rating_date')[:5]:
        recent_activity.append({
            'description': f'Оценил трек "{rating.track.name}" на {rating.value} звезд',
            'timestamp': rating.rating_date
        })

    context = {
        'user_playlists': user_playlists,
        'recent_playlists': recent_playlists,
        'user_ratings': user_ratings,
        'playlists_count': playlists_count,
        'total_tracks': total_tracks,
        'total_duration': total_duration,
        'total_play_count': total_play_count,
        'recent_activity': recent_activity,
    }
    return render(request, 'music/profile.html', context)


@login_required
@require_POST
def toggle_admin_role(request):
    """Переключение роли пользователя между обычным и администратором"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Не авторизован'}, status=401)
    
    print(f"Toggle admin role request from user: {request.user.login}, current role: {request.user.role}")
    
    try:
        # Переключаем роль
        if request.user.role == 'admin':
            request.user.role = 'user'
            message = 'Роль изменена на пользователь'
        else:
            request.user.role = 'admin'
            message = 'Роль изменена на администратор'
        
        request.user.save()
        
        return JsonResponse({
            'success': True,
            'message': message,
            'new_role': request.user.role
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# API представления для AJAX
@csrf_exempt
@require_POST
def api_rate_track(request, track_id):
    """API для оценки трека"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Не авторизован'}, status=401)
    
    try:
        data = json.loads(request.body)
        rating_value = int(data.get('rating', 0))
        
        if not (1 <= rating_value <= 5):
            return JsonResponse({'error': 'Оценка должна быть от 1 до 5'}, status=400)
        
        track = get_object_or_404(Track, pk=track_id)
        rating, created = TrackRating.objects.get_or_create(
                user=request.user,
            track=track,
            defaults={'value': rating_value}
        )
        
        if not created:
            rating.value = rating_value
            rating.save()
        
        return JsonResponse({'success': True, 'rating': rating_value})
    
    except (ValueError, json.JSONDecodeError):
        return JsonResponse({'error': 'Неверные данные'}, status=400)


@csrf_exempt
@require_POST
def api_rate_album(request, album_id):
    """API для оценки альбома"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Не авторизован'}, status=401)
    
    try:
        data = json.loads(request.body)
        rating_value = int(data.get('rating', 0))
        
        if not (1 <= rating_value <= 5):
            return JsonResponse({'error': 'Оценка должна быть от 1 до 5'}, status=400)
        
        album = get_object_or_404(Album, pk=album_id)
        rating, created = AlbumRating.objects.get_or_create(
            user=request.user,
            album=album,
            defaults={'value': rating_value}
        )
        
        if not created:
            rating.value = rating_value
            rating.save()
        
        return JsonResponse({'success': True, 'rating': rating_value})
    
    except (ValueError, json.JSONDecodeError):
        return JsonResponse({'error': 'Неверные данные'}, status=400)


@csrf_exempt
@require_POST
def api_add_comment(request, track_id):
    """API для добавления комментария"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Не авторизован'}, status=401)
    
    try:
        data = json.loads(request.body)
        text = data.get('text', '').strip()
        
        if not text:
            return JsonResponse({'error': 'Текст комментария не может быть пустым'}, status=400)
        
        track = get_object_or_404(Track, pk=track_id)
        comment = Comment.objects.create(
                user=request.user,
            track=track,
            text=text
        )
        
        return JsonResponse({
            'success': True,
            'comment': {
                'id': comment.id,
                'text': comment.text,
                'user': comment.user.login,
                'created_at': comment.created_at.isoformat()
            }
        })
    
    except (ValueError, json.JSONDecodeError):
        return JsonResponse({'error': 'Неверные данные'}, status=400)


@csrf_exempt
@require_POST
def api_delete_comment(request, comment_id):
    """API для удаления комментария"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Не авторизован'}, status=401)
    
    try:
        comment = get_object_or_404(Comment, pk=comment_id, user=request.user)
        comment.delete()
        
        return JsonResponse({'success': True})
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def api_get_playlists(request):
    """API для получения плейлистов пользователя"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Не авторизован'}, status=401)
    
    try:
        playlists = Playlist.objects.filter(user=request.user).values('id', 'name')
        return JsonResponse({'playlists': list(playlists)})
    
    except Exception as e:
        print(f"Error in api_get_playlists: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


def api_play_track(request, track_id):
    """API для воспроизведения трека"""
    try:
        track = get_object_or_404(Track, pk=track_id)
        
        # Увеличиваем счетчик прослушиваний
        track.play_count += 1
        track.save()
        
        # Возвращаем URL файла для воспроизведения
        if track.file:
            return JsonResponse({
                'success': True,
                'file_url': track.file.url,
                'track_name': track.name
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Файл трека недоступен'
            })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@csrf_exempt
@require_POST
def api_add_track_to_playlist(request, playlist_id):
    """API для добавления трека в плейлист"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Не авторизован'}, status=401)
    
    try:
        data = json.loads(request.body)
        track_id = data.get('track_id')
        
        if not track_id:
            return JsonResponse({'error': 'ID трека не указан'}, status=400)
        
        playlist = get_object_or_404(Playlist, pk=playlist_id, user=request.user)
        track = get_object_or_404(Track, pk=track_id)
        
        # Проверяем, не добавлен ли уже трек
        if PlaylistTrack.objects.filter(playlist=playlist, track=track).exists():
            return JsonResponse({'error': 'Трек уже добавлен в этот плейлист'}, status=400)
        
        PlaylistTrack.objects.create(playlist=playlist, track=track)
        
        return JsonResponse({'success': True})
    
    except (ValueError, json.JSONDecodeError):
        return JsonResponse({'error': 'Неверные данные'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_POST
def api_remove_track_from_playlist(request, playlist_id):
    """API для удаления трека из плейлиста (AJAX)"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Не авторизован'}, status=401)
    try:
        data = json.loads(request.body)
        track_id = data.get('track_id')
        if not track_id:
            return JsonResponse({'error': 'ID трека не указан'}, status=400)
        playlist = get_object_or_404(Playlist, pk=playlist_id, user=request.user)
        track = get_object_or_404(Track, pk=track_id)
        if not playlist.tracks.filter(pk=track.pk).exists():
            return JsonResponse({'error': 'Трек не найден в плейлисте'}, status=404)
        playlist.tracks.remove(track)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# Админ панель представления
@login_required
def admin_panel(request):
    """Главная страница админ панели"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    # Статистика
    total_tracks = Track.objects.count()
    total_albums = Album.objects.count()
    total_artists = Artist.objects.count()
    total_groups = Group.objects.count()
    
    context = {
        'total_tracks': total_tracks,
        'total_albums': total_albums,
        'total_artists': total_artists,
        'total_groups': total_groups,
    }
    return render(request, 'music/admin/admin_panel.html', context)


@login_required
def admin_send_email(request):
    """Отправка email-уведомлений пользователям (только для админа)"""
    if not request.user.role == 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:admin_panel')

    users = User.objects.all().values('id', 'login', 'email')

    if request.method == 'POST':
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()
        target = request.POST.get('target', 'all')

        if not subject or not message:
            messages.error(request, 'Тема и сообщение обязательны')
            return redirect('music:admin_send_email')

        recipients = []
        if target == 'all':
            recipients = [u['email'] for u in users if u['email']]
        elif target.startswith('user:'):
            uid = target.split(':', 1)[1]
            user = User.objects.filter(pk=uid).first()
            if user and user.email:
                recipients = [user.email]

        # Отправляем письма через EmailMessage
        try:
            sent_count = 0
            failed_emails = []
            
            for email in recipients:
                try:
                    msg = EmailMessage(
                        subject=subject, 
                        body=message, 
                        from_email='yakov.goryachev@mail.ru',
                        to=[email]
                    )
                    msg.send(fail_silently=False)
                    sent_count += 1
                    print(f"Email sent successfully to: {email}")
                except Exception as email_error:
                    print(f"Failed to send email to {email}: {str(email_error)}")
                    failed_emails.append(email)
                    continue

            if sent_count > 0:
                success_msg = f'Успешно отправлено {sent_count} писем'
                if failed_emails:
                    success_msg += f'. Не удалось отправить {len(failed_emails)} писем'
                messages.success(request, success_msg)
            else:
                messages.error(request, 'Не удалось отправить ни одного письма')
                
            return redirect('music:admin_panel')
            
        except Exception as e:
            print(f"Email sending error: {str(e)}")
            messages.error(request, f'Ошибка отправки: {str(e)}')
            return redirect('music:admin_send_email')

    context = {
        'users': users,
    }
    return render(request, 'music/admin/admin_send_email.html', context)


@login_required
def admin_reports(request):
    if not request.user.role == 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:admin_panel')

    return render(request, 'music/admin/admin_reports.html')


@login_required
def admin_generate_report(request):
    """Генерация расширенных отчетов с аналитикой"""
    if not request.user.role == 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:admin_panel')

    fmt = request.GET.get('format', 'xlsx')
    
    # Импорты для генерации отчетов
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Требуется пакет openpyxl для генерации Excel. Установите его: pip install openpyxl')
        return redirect('music:admin_reports')
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.fonts import addMapping
    except ImportError:
        messages.error(request, 'Требуется пакет reportlab для генерации PDF. Установите его: pip install reportlab')
        return redirect('music:admin_reports')

    # === ОСНОВНАЯ СТАТИСТИКА ===
    total_users = User.objects.count()
    total_tracks = Track.objects.count()
    total_albums = Album.objects.count()
    total_artists = Artist.objects.count()
    total_groups = Group.objects.count()
    total_playlists = Playlist.objects.count()
    total_comments = Comment.objects.count()
    total_track_ratings = TrackRating.objects.count()
    total_album_ratings = AlbumRating.objects.count()
    total_genres = Genre.objects.count()

    # === ПОПУЛЯРНОСТЬ ===
    top_tracks = Track.objects.annotate(cnt=Sum('play_count')).order_by('-cnt')[:20]
    # Для альбомов считаем сумму прослушиваний всех треков
    top_albums = Album.objects.annotate(
        total_plays=Sum('track__play_count')
    ).filter(total_plays__isnull=False).order_by('-total_plays')[:20]
    top_artists_by_albums = Artist.objects.annotate(albums_count=Count('album')).order_by('-albums_count')[:20]
    top_groups_by_albums = Group.objects.annotate(albums_count=Count('album')).order_by('-albums_count')[:20]

    # === РЕЙТИНГИ ===
    best_tracks = Track.objects.annotate(avg=Avg('ratings__value')).filter(avg__isnull=False).order_by('-avg')[:20]
    worst_tracks = Track.objects.annotate(avg=Avg('ratings__value')).filter(avg__isnull=False).order_by('avg')[:20]
    best_albums = Album.objects.annotate(avg=Avg('ratings__value')).filter(avg__isnull=False).order_by('-avg')[:20]

    # === ЖАНРОВАЯ АНАЛИТИКА ===
    genre_stats = Genre.objects.annotate(
        tracks_count=Count('track'),
        avg_rating=Avg('track__ratings__value')
    ).order_by('-tracks_count')[:15]

    # === ПОЛЬЗОВАТЕЛЬСКАЯ АКТИВНОСТЬ ===
    most_active_users = User.objects.annotate(
        playlists_count=Count('playlists'),
        comments_count=Count('comment'),
        ratings_count=Count('trackrating') + Count('albumrating')
    ).order_by('-playlists_count')[:15]

    # === ВРЕМЕННАЯ АНАЛИТИКА ===
    from django.utils import timezone
    from datetime import timedelta
    
    # Новые пользователи за последние 30 дней
    thirty_days_ago = timezone.now() - timedelta(days=30)
    new_users_30d = User.objects.filter(registration_date__gte=thirty_days_ago).count()
    
    # Новые плейлисты за последние 30 дней
    new_playlists_30d = Playlist.objects.filter(creation_date__gte=thirty_days_ago).count()
    
    # Новые комментарии за последние 30 дней
    new_comments_30d = Comment.objects.filter(created_at__gte=thirty_days_ago).count()

    # === САМЫЕ ДЛИННЫЕ И КОРОТКИЕ ТРЕКИ ===
    longest_tracks = Track.objects.filter(duration__isnull=False).order_by('-duration')[:10]
    shortest_tracks = Track.objects.filter(duration__isnull=False).order_by('duration')[:10]

    # === АЛЬБОМЫ ПО ГОДАМ ===
    # Для SQLite используем strftime вместо EXTRACT
    albums_by_year = Album.objects.filter(release_date__isnull=False).extra(
        select={'year': "strftime('%Y', release_date)"}
    ).values('year').annotate(count=Count('id')).order_by('-year')[:10]

    # === САМЫЕ КОММЕНТИРУЕМЫЕ ТРЕКИ ===
    most_commented_tracks = Track.objects.annotate(
        comments_count=Count('comment')
    ).filter(comments_count__gt=0).order_by('-comments_count')[:15]

    # === СТАТИСТИКА ПО РОЛЯМ ПОЛЬЗОВАТЕЛЕЙ ===
    users_by_role = User.objects.values('role').annotate(count=Count('id')).order_by('-count')

    # === САМЫЕ ПОПУЛЯРНЫЕ ПЛЕЙЛИСТЫ ===
    popular_playlists = Playlist.objects.annotate(
        tracks_count=Count('tracks')
    ).order_by('-tracks_count')[:15]

    if fmt == 'xlsx':

        wb = openpyxl.Workbook()
        
        # === ОСНОВНАЯ СТАТИСТИКА ===
        ws = wb.active
        ws.title = 'Общая статистика'
        ws.append(['Метрика', 'Значение'])
        ws.append(['Пользователи', total_users])
        ws.append(['Треки', total_tracks])
        ws.append(['Альбомы', total_albums])
        ws.append(['Артисты', total_artists])
        ws.append(['Группы', total_groups])
        ws.append(['Плейлисты', total_playlists])
        ws.append(['Комментарии', total_comments])
        ws.append(['Оценки треков', total_track_ratings])
        ws.append(['Оценки альбомов', total_album_ratings])
        ws.append(['Жанры', total_genres])
        ws.append(['', ''])
        ws.append(['=== АКТИВНОСТЬ ЗА 30 ДНЕЙ ===', ''])
        ws.append(['Новые пользователи', new_users_30d])
        ws.append(['Новые плейлисты', new_playlists_30d])
        ws.append(['Новые комментарии', new_comments_30d])

        # === ТОП ТРЕКОВ ПО ПРОСЛУШИВАНИЯМ ===
        ws2 = wb.create_sheet('Топ треков по прослушиваниям')
        ws2.append(['#', 'Название трека', 'Прослушивания', 'Альбом'])
        for i, t in enumerate(top_tracks, start=1):
            album_name = t.album.name if t.album else 'Без альбома'
            ws2.append([i, t.name, t.play_count or 0, album_name])

        # === ТОП АЛЬБОМОВ ПО ПРОСЛУШИВАНИЯМ ===
        ws3 = wb.create_sheet('Топ альбомов по прослушиваниям')
        ws3.append(['#', 'Название альбома', 'Прослушивания', 'Группа/Артист'])
        for i, a in enumerate(top_albums, start=1):
            artist_name = a.group.name if a.group else (a.artist.name if a.artist else 'Неизвестно')
            ws3.append([i, a.name, a.total_plays or 0, artist_name])

        # === ТОП АРТИСТОВ ПО КОЛИЧЕСТВУ АЛЬБОМОВ ===
        ws4 = wb.create_sheet('Топ артистов по альбомам')
        ws4.append(['#', 'Имя артиста', 'Количество альбомов'])
        for i, a in enumerate(top_artists_by_albums, start=1):
            ws4.append([i, a.name, a.albums_count])

        # === ТОП ГРУПП ПО КОЛИЧЕСТВУ АЛЬБОМОВ ===
        ws5 = wb.create_sheet('Топ групп по альбомам')
        ws5.append(['#', 'Название группы', 'Количество альбомов'])
        for i, g in enumerate(top_groups_by_albums, start=1):
            ws5.append([i, g.name, g.albums_count])

        # === ЛУЧШИЕ ТРЕКИ ПО РЕЙТИНГУ ===
        ws6 = wb.create_sheet('Лучшие треки по рейтингу')
        ws6.append(['#', 'Название трека', 'Средний рейтинг', 'Количество оценок'])
        for i, t in enumerate(best_tracks, start=1):
            rating_count = t.ratings.count()
            ws6.append([i, t.name, round(t.avg, 2), rating_count])

        # === ХУДШИЕ ТРЕКИ ПО РЕЙТИНГУ ===
        ws7 = wb.create_sheet('Худшие треки по рейтингу')
        ws7.append(['#', 'Название трека', 'Средний рейтинг', 'Количество оценок'])
        for i, t in enumerate(worst_tracks, start=1):
            rating_count = t.ratings.count()
            ws7.append([i, t.name, round(t.avg, 2), rating_count])

        # === ЛУЧШИЕ АЛЬБОМЫ ПО РЕЙТИНГУ ===
        ws8 = wb.create_sheet('Лучшие альбомы по рейтингу')
        ws8.append(['#', 'Название альбома', 'Средний рейтинг', 'Количество оценок'])
        for i, a in enumerate(best_albums, start=1):
            rating_count = a.ratings.count()
            ws8.append([i, a.name, round(a.avg, 2), rating_count])

        # === ЖАНРОВАЯ СТАТИСТИКА ===
        ws9 = wb.create_sheet('Статистика по жанрам')
        ws9.append(['#', 'Жанр', 'Количество треков', 'Средний рейтинг'])
        for i, g in enumerate(genre_stats, start=1):
            avg_rating = round(g.avg_rating, 2) if g.avg_rating else 0
            ws9.append([i, g.name, g.tracks_count, avg_rating])

        # === САМЫЕ АКТИВНЫЕ ПОЛЬЗОВАТЕЛИ ===
        ws10 = wb.create_sheet('Самые активные пользователи')
        ws10.append(['#', 'Пользователь', 'Плейлисты', 'Комментарии', 'Оценки'])
        for i, u in enumerate(most_active_users, start=1):
            ws10.append([i, u.login, u.playlists_count, u.comments_count, u.ratings_count])

        # === САМЫЕ ДЛИННЫЕ ТРЕКИ ===
        ws11 = wb.create_sheet('Самые длинные треки')
        ws11.append(['#', 'Название трека', 'Длительность (мин:сек)', 'Альбом'])
        for i, t in enumerate(longest_tracks, start=1):
            duration_min = t.duration // 60
            duration_sec = t.duration % 60
            duration_str = f"{duration_min}:{duration_sec:02d}"
            album_name = t.album.name if t.album else 'Без альбома'
            ws11.append([i, t.name, duration_str, album_name])

        # === САМЫЕ КОРОТКИЕ ТРЕКИ ===
        ws12 = wb.create_sheet('Самые короткие треки')
        ws12.append(['#', 'Название трека', 'Длительность (мин:сек)', 'Альбом'])
        for i, t in enumerate(shortest_tracks, start=1):
            duration_min = t.duration // 60
            duration_sec = t.duration % 60
            duration_str = f"{duration_min}:{duration_sec:02d}"
            album_name = t.album.name if t.album else 'Без альбома'
            ws12.append([i, t.name, duration_str, album_name])

        # === АЛЬБОМЫ ПО ГОДАМ ===
        ws13 = wb.create_sheet('Альбомы по годам')
        ws13.append(['Год', 'Количество альбомов'])
        for year_data in albums_by_year:
            ws13.append([int(year_data['year']), year_data['count']])

        # === САМЫЕ КОММЕНТИРУЕМЫЕ ТРЕКИ ===
        ws14 = wb.create_sheet('Самые комментируемые треки')
        ws14.append(['#', 'Название трека', 'Количество комментариев', 'Альбом'])
        for i, t in enumerate(most_commented_tracks, start=1):
            album_name = t.album.name if t.album else 'Без альбома'
            ws14.append([i, t.name, t.comments_count, album_name])

        # === СТАТИСТИКА ПО РОЛЯМ ===
        ws15 = wb.create_sheet('Пользователи по ролям')
        ws15.append(['Роль', 'Количество'])
        for role_data in users_by_role:
            role_name = {'user': 'Пользователь', 'admin': 'Администратор', 'moderator': 'Модератор'}.get(role_data['role'], role_data['role'])
            ws15.append([role_name, role_data['count']])

        # === ПОПУЛЯРНЫЕ ПЛЕЙЛИСТЫ ===
        ws16 = wb.create_sheet('Популярные плейлисты')
        ws16.append(['#', 'Название плейлиста', 'Количество треков', 'Владелец', 'Публичный'])
        for i, p in enumerate(popular_playlists, start=1):
            is_public = 'Да' if p.is_public else 'Нет'
            ws16.append([i, p.name, p.tracks_count, p.user.login, is_public])

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        response = HttpResponse(out.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="music_advanced_report.xlsx"'
        return response

    elif fmt == 'pdf':

        # Настройка шрифтов для поддержки кириллицы
        try:
            # Попробуем использовать системный шрифт Arial
            pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
            pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))
            font_name = 'Arial'
            font_bold = 'Arial-Bold'
        except:
            # Fallback на встроенные шрифты ReportLab с поддержкой Unicode
            font_name = 'Helvetica'
            font_bold = 'Helvetica-Bold'
        
        # Функция для безопасного отображения текста
        def safe_text(text):
            if text is None:
                return ""
            # Заменяем проблемные символы на безопасные
            return str(text).encode('utf-8', 'replace').decode('utf-8', 'replace')

        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 40
        c.setFont(font_bold, 14)
        c.drawString(40, y, 'Music Service - Report')
        y -= 30
        # Summary
        c.setFont(font_bold, 12)
        c.drawString(40, y, 'Общая статистика')
        y -= 20
        c.setFont(font_name, 10)
        for label, value in [
            ('Пользователи', total_users),
            ('Треки', total_tracks),
            ('Альбомы', total_albums),
            ('Артисты', total_artists),
            ('Группы', total_groups),
            ('Плейлисты', total_playlists),
            ('Комментарии', total_comments),
            ('Оценки треков', total_track_ratings),
            ('Оценки альбомов', total_album_ratings),
            ('Жанры', total_genres),
        ]:
            c.drawString(40, y, f"{label}: {value}")
            y -= 14
            if y < 80:
                c.showPage()
                y = height - 40

        # Активность за 30 дней
        if y < 100:
            c.showPage()
            y = height - 40
        c.setFont(font_bold, 12)
        c.drawString(40, y, 'Активность за 30 дней')
        y -= 20
        c.setFont(font_name, 10)
        for label, value in [
            ('Новые пользователи', new_users_30d),
            ('Новые плейлисты', new_playlists_30d),
            ('Новые комментарии', new_comments_30d),
        ]:
            c.drawString(40, y, f"{label}: {value}")
            y -= 14
            if y < 80:
                c.showPage()
                y = height - 40

        # Top Tracks by plays
        if y < 140:
            c.showPage()
            y = height - 40
        c.setFont(font_bold, 12)
        c.drawString(40, y, 'Топ треков по прослушиваниям')
        y -= 20
        c.setFont(font_name, 10)
        for i, t in enumerate(top_tracks[:15], start=1):
            line = f"{i}. {safe_text(t.name)} — {t.play_count or 0} прослушиваний"
            c.drawString(40, y, line)
            y -= 14
            if y < 80:
                c.showPage()
                y = height - 40

        # Best Tracks by rating
        if y < 140:
            c.showPage()
            y = height - 40
        c.setFont(font_bold, 12)
        c.drawString(40, y, 'Лучшие треки по рейтингу')
        y -= 20
        c.setFont(font_name, 10)
        for i, t in enumerate(best_tracks[:15], start=1):
            rating_count = t.ratings.count()
            line = f"{i}. {safe_text(t.name)} — {round(t.avg, 2)} ({rating_count} оценок)"
            c.drawString(40, y, line)
            y -= 14
            if y < 80:
                c.showPage()
                y = height - 40

        # Top Albums by plays
        if y < 140:
            c.showPage()
            y = height - 40
        c.setFont(font_bold, 12)
        c.drawString(40, y, 'Топ альбомов по прослушиваниям')
        y -= 20
        c.setFont(font_name, 10)
        for i, a in enumerate(top_albums[:15], start=1):
            line = f"{i}. {safe_text(a.name)} — {a.total_plays or 0} прослушиваний"
            c.drawString(40, y, line)
            y -= 14
            if y < 80:
                c.showPage()
                y = height - 40

        # Genre Statistics
        if y < 140:
            c.showPage()
            y = height - 40
        c.setFont(font_bold, 12)
        c.drawString(40, y, 'Статистика по жанрам')
        y -= 20
        c.setFont(font_name, 10)
        for i, g in enumerate(genre_stats[:10], start=1):
            avg_rating = round(g.avg_rating, 2) if g.avg_rating else 0
            line = f"{i}. {safe_text(g.name)} — {g.tracks_count} треков, рейтинг: {avg_rating}"
            c.drawString(40, y, line)
            y -= 14
            if y < 80:
                c.showPage()
                y = height - 40

        # Most Active Users
        if y < 140:
            c.showPage()
            y = height - 40
        c.setFont(font_bold, 12)
        c.drawString(40, y, 'Самые активные пользователи')
        y -= 20
        c.setFont(font_name, 10)
        for i, u in enumerate(most_active_users[:10], start=1):
            line = f"{i}. {safe_text(u.login)} — {u.playlists_count} плейлистов, {u.comments_count} комментариев"
            c.drawString(40, y, line)
            y -= 14
            if y < 80:
                c.showPage()
                y = height - 40

        # Longest Tracks
        if y < 140:
            c.showPage()
            y = height - 40
        c.setFont(font_bold, 12)
        c.drawString(40, y, 'Самые длинные треки')
        y -= 20
        c.setFont(font_name, 10)
        for i, t in enumerate(longest_tracks[:10], start=1):
            duration_min = t.duration // 60
            duration_sec = t.duration % 60
            duration_str = f"{duration_min}:{duration_sec:02d}"
            line = f"{i}. {safe_text(t.name)} — {duration_str}"
            c.drawString(40, y, line)
            y -= 14
            if y < 80:
                c.showPage()
                y = height - 40

        # Most Commented Tracks
        if y < 140:
            c.showPage()
            y = height - 40
        c.setFont(font_bold, 12)
        c.drawString(40, y, 'Самые комментируемые треки')
        y -= 20
        c.setFont(font_name, 10)
        for i, t in enumerate(most_commented_tracks[:10], start=1):
            line = f"{i}. {safe_text(t.name)} — {t.comments_count} комментариев"
            c.drawString(40, y, line)
            y -= 14
            if y < 80:
                c.showPage()
                y = height - 40

        c.showPage()
        c.save()
        buffer.seek(0)
        
        # Создаем HttpResponse с правильными заголовками для скачивания
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="music_service_report.pdf"'
        return response

    else:
        messages.error(request, 'Неподдерживаемый формат')
        return redirect('music:admin_reports')


@login_required
def admin_tracks(request):
    """Управление треками"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    query = request.GET.get('q', '')
    genre_filter = request.GET.get('genre', '')
    
    tracks = Track.objects.select_related('album', 'album__artist', 'album__group').prefetch_related('genres')
    
    if query:
        tracks = tracks.filter(
            Q(name__icontains=query) |
            Q(album__name__icontains=query) |
            Q(album__artist__name__icontains=query) |
            Q(album__group__name__icontains=query) |
            Q(genres__name__icontains=query)
        ).distinct()
    
    if genre_filter:
        tracks = tracks.filter(genres__name__icontains=genre_filter)
    
    tracks = tracks.order_by('-id')
    
    genres = Genre.objects.all().order_by('name')
    
    context = {
        'tracks': tracks,
        'query': query,
        'genre_filter': genre_filter,
        'genres': genres,
    }
    return render(request, 'music/admin/admin_tracks.html', context)


@login_required
def admin_create_track(request):
    """Создание нового трека"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    if request.method == 'POST':
        form = TrackCreateForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Получаем данные из формы
                name = form.cleaned_data['name']
                file = form.cleaned_data['file']
                duration = form.cleaned_data.get('duration')
                artist_name = form.cleaned_data.get('artist_name')
                group_name = form.cleaned_data.get('group_name')
                album_name = form.cleaned_data.get('album_name')
                album_release_date = form.cleaned_data.get('album_release_date')
                genre_names = form.cleaned_data.get('genre_names', [])
                
                # Создаем или получаем артиста/группу
                artist = None
                group = None
                
                if group_name:
                    group, created = Group.objects.get_or_create(name=group_name)
                    if created:
                        messages.info(request, f'Создана новая группа: {group_name}')
                
                if artist_name:
                    artist, created = Artist.objects.get_or_create(name=artist_name)
                    if created:
                        messages.info(request, f'Создан новый артист: {artist_name}')
                
                # Создаем или получаем альбом
                album = None
                if album_name:
                    album, created = Album.objects.get_or_create(
                        name=album_name,
                        defaults={
                            'artist': artist,
                            'group': group,
                            'release_date': album_release_date
                        }
                    )
                    if created:
                        messages.info(request, f'Создан новый альбом: {album_name}')
                
                # Создаем трек
                track = Track.objects.create(
                    name=name,
                    file=file,
                    photo=form.cleaned_data.get('photo'),
                    album=album,
                    duration=duration
                )
                
                # Создаем или получаем жанры
                for genre_name in genre_names:
                    if genre_name:
                        genre, created = Genre.objects.get_or_create(name=genre_name)
                        if created:
                            messages.info(request, f'Создан новый жанр: {genre_name}')
                        TrackGenre.objects.create(track=track, genre=genre)
                
                messages.success(request, f'Трек "{name}" успешно создан!')
                return redirect('music:admin_tracks')
                
            except Exception as e:
                messages.error(request, f'Ошибка создания трека: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = TrackCreateForm()
    
    context = {
        'form': form,
    }
    return render(request, 'music/admin/admin_create_track.html', context)


@login_required
def admin_edit_track(request, pk):
    """Редактирование трека"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    track = get_object_or_404(Track, pk=pk)
    
    if request.method == 'POST':
        # Обработка редактирования трека
        name = request.POST.get('name')
        artist_name = request.POST.get('artist_name')
        group_name = request.POST.get('group_name')
        album_name = request.POST.get('album_name')
        album_release_date = request.POST.get('album_release_date')
        duration = request.POST.get('duration')
        genre_ids = request.POST.getlist('genres')
        
        if name:
            try:
                # Обновляем основные поля
                track.name = name
                
                if duration:
                    track.duration = int(duration)
                else:
                    track.duration = None
                
                # Обработка загрузки нового файла
                if 'file' in request.FILES:
                    track.file = request.FILES['file']
                
                # Обработка загрузки нового фото
                if 'photo' in request.FILES:
                    track.photo = request.FILES['photo']
                
                # Обработка артиста
                artist = None
                if artist_name:
                    artist, created = Artist.objects.get_or_create(
                        name=artist_name,
                        defaults={'biography': ''}
                    )
                
                # Обработка группы
                group = None
                if group_name:
                    group, created = Group.objects.get_or_create(
                        name=group_name,
                        defaults={'description': ''}
                    )
                
                # Обработка альбома
                album = None
                if album_name:
                    album, created = Album.objects.get_or_create(
                        name=album_name,
                        defaults={
                            'release_date': album_release_date if album_release_date else None,
                            'artist': artist,
                            'group': group,
                        }
                    )
                    # Обновляем альбом если он уже существует
                    if not created:
                        album.artist = artist
                        album.group = group
                        if album_release_date:
                            album.release_date = album_release_date
                        album.save()
                
                track.album = album
                track.save()
                
                # Обновляем жанры
                track.genres.clear()
                for genre_id in genre_ids:
                    if genre_id:
                        genre = Genre.objects.get(pk=genre_id)
                        TrackGenre.objects.get_or_create(track=track, genre=genre)
                
                messages.success(request, f'Трек "{name}" успешно обновлен!')
                return redirect('music:admin_tracks')
            except Exception as e:
                messages.error(request, f'Ошибка обновления трека: {str(e)}')
        else:
            messages.error(request, 'Название трека обязательно!')
    
    albums = Album.objects.all()
    genres = Genre.objects.all()
    selected_genres = [genre.id for genre in track.genres.all()]
    
    context = {
        'track': track,
        'albums': albums,
        'genres': genres,
        'selected_genres': selected_genres,
    }
    return render(request, 'music/admin/admin_edit_track.html', context)


@login_required
def admin_delete_track(request, pk):
    """Удаление трека"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    track = get_object_or_404(Track, pk=pk)
    
    if request.method == 'POST':
        track_name = track.name
        track.delete()
        messages.success(request, f'Трек "{track_name}" успешно удален!')
        return redirect('music:admin_tracks')
    
    context = {
        'track': track,
    }
    return render(request, 'music/admin/admin_delete_track.html', context)


@login_required
def admin_albums(request):
    """Управление альбомами"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    albums = Album.objects.select_related('artist', 'group').order_by('-id')
    
    context = {
        'albums': albums,
    }
    return render(request, 'music/admin/admin_albums.html', context)


@login_required
def admin_edit_album(request, album_id):
    """Редактирование альбома"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    try:
        album = Album.objects.get(id=album_id)
    except Album.DoesNotExist:
        messages.error(request, 'Альбом не найден!')
        return redirect('music:admin_albums')
    
    if request.method == 'POST':
        # Обработка редактирования альбома
        name = request.POST.get('name')
        artist_id = request.POST.get('artist')
        group_id = request.POST.get('group')
        release_date = request.POST.get('release_date')
        photo = request.FILES.get('photo')
        
        if name and (artist_id or group_id):
            try:
                artist = None
                group = None
                if artist_id:
                    artist = Artist.objects.get(pk=artist_id)
                if group_id:
                    group = Group.objects.get(pk=group_id)
                album.name = name
                album.artist = artist
                album.group = group
                if release_date:
                    album.release_date = release_date
                # Обновляем фото только если загружено новое
                if photo:
                    album.photo = photo
                album.save()
                messages.success(request, f'Альбом "{name}" успешно обновлен!')
                return redirect('music:admin_albums')
            except Exception as e:
                messages.error(request, f'Ошибка обновления альбома: {str(e)}')
        else:
            messages.error(request, 'Заполните все обязательные поля!')
    
    artists = Artist.objects.all()
    groups = Group.objects.all()
    
    context = {
        'album': album,
        'artists': artists,
        'groups': groups,
    }
    return render(request, 'music/admin/admin_edit_album.html', context)


@login_required
def admin_delete_album(request, album_id):
    """Удаление альбома"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    try:
        album = Album.objects.get(id=album_id)
        album_name = album.name
        album.delete()
        messages.success(request, f'Альбом "{album_name}" успешно удален!')
    except Album.DoesNotExist:
        messages.error(request, 'Альбом не найден!')
    except Exception as e:
        messages.error(request, f'Ошибка удаления альбома: {str(e)}')
    
    return redirect('music:admin_albums')


@login_required
def admin_create_album(request):
    """Создание нового альбома"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    if request.method == 'POST':
        # Обработка создания альбома
        name = request.POST.get('name')
        artist_id = request.POST.get('artist')
        group_id = request.POST.get('group')
        release_date = request.POST.get('release_date')
        photo = request.FILES.get('photo')
        
        if name and (artist_id or group_id):
            try:
                artist = None
                group = None
                
                if artist_id:
                    artist = Artist.objects.get(pk=artist_id)
                if group_id:
                    group = Group.objects.get(pk=group_id)
                
                album = Album.objects.create(
                    name=name,
                    artist=artist,
                    group=group,
                    release_date=release_date if release_date else None,
                    photo=photo if photo else None
                )
                
                messages.success(request, f'Альбом "{name}" успешно создан!')
                return redirect('music:admin_albums')
            except Exception as e:
                messages.error(request, f'Ошибка создания альбома: {str(e)}')
        else:
            messages.error(request, 'Заполните все обязательные поля!')
    
    artists = Artist.objects.all()
    groups = Group.objects.all()
    
    context = {
        'artists': artists,
        'groups': groups,
    }
    return render(request, 'music/admin/admin_create_album.html', context)


@login_required
def admin_artists(request):
    """Управление артистами"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    artists = Artist.objects.all().order_by('-id')
    
    context = {
        'artists': artists,
    }
    return render(request, 'music/admin/admin_artists.html', context)


@login_required
def admin_create_artist(request):
    """Создание нового артиста"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    if request.method == 'POST':
        # Обработка создания артиста
        name = request.POST.get('name')
        bio = request.POST.get('bio')
        artist_role = request.POST.get('artist_role')
        avatar = request.FILES.get('avatar')
        
        if name:
            try:
                artist = Artist.objects.create(
                    name=name,
                    biography=bio if bio else '',
                    artist_role=artist_role if artist_role else '',
                    avatar=avatar if avatar else None
                )
                
                messages.success(request, f'Артист "{name}" успешно создан!')
                return redirect('music:admin_artists')
            except Exception as e:
                messages.error(request, f'Ошибка создания артиста: {str(e)}')
        else:
            messages.error(request, 'Заполните все обязательные поля!')
    
    context = {}
    return render(request, 'music/admin/admin_create_artist.html', context)


@login_required
def admin_edit_artist(request, artist_id):
    """Редактирование артиста"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    try:
        artist = Artist.objects.get(id=artist_id)
    except Artist.DoesNotExist:
        messages.error(request, 'Артист не найден!')
        return redirect('music:admin_artists')
    
    if request.method == 'POST':
        # Обработка редактирования артиста
        name = request.POST.get('name')
        bio = request.POST.get('bio')
        artist_role = request.POST.get('artist_role')
        avatar = request.FILES.get('avatar')
        
        if name:
            try:
                artist.name = name
                artist.biography = bio if bio else ''
                artist.artist_role = artist_role if artist_role else ''
                
                # Обновляем фото только если загружено новое
                if avatar:
                    artist.avatar = avatar
                
                artist.save()
                
                messages.success(request, f'Артист "{name}" успешно обновлен!')
                return redirect('music:admin_artists')
            except Exception as e:
                messages.error(request, f'Ошибка обновления артиста: {str(e)}')
        else:
            messages.error(request, 'Заполните все обязательные поля!')
    
    context = {
        'artist': artist,
    }
    return render(request, 'music/admin/admin_edit_artist.html', context)


@login_required
def admin_delete_artist(request, artist_id):
    """Удаление артиста"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    try:
        artist = Artist.objects.get(id=artist_id)
        artist_name = artist.name
        artist.delete()
        messages.success(request, f'Артист "{artist_name}" успешно удален!')
    except Artist.DoesNotExist:
        messages.error(request, 'Артист не найден!')
    except Exception as e:
        messages.error(request, f'Ошибка удаления артиста: {str(e)}')
    
    return redirect('music:admin_artists')
@login_required
def admin_groups(request):
    """Управление группами"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    groups = Group.objects.all().order_by('-id')
    
    context = {
        'groups': groups,
    }
    return render(request, 'music/admin/admin_groups.html', context)


@login_required
def admin_edit_group(request, group_id):
    """Редактирование группы"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    try:
        group = Group.objects.get(id=group_id)
    except Group.DoesNotExist:
        messages.error(request, 'Группа не найдена!')
        return redirect('music:admin_groups')
    
    if request.method == 'POST':
        # Обработка редактирования группы
        name = request.POST.get('name')
        description = request.POST.get('description')
        photo = request.FILES.get('photo')
        # Динамические участники (массивы синхронные)
        participant_artists = request.POST.getlist('participant_artist')
        participant_roles = request.POST.getlist('participant_role')
        
        if name:
            try:
                group.name = name
                group.description = description if description else ''
                
                # Обновляем фото только если загружено новое
                if photo:
                    group.photo = photo
                
                group.save()

                # Обновляем участников группы через ArtistGroup
                ArtistGroup.objects.filter(group=group).delete()
                for aid, role in zip(participant_artists, participant_roles):
                    if not aid:
                        continue
                    try:
                        artist_obj = Artist.objects.get(pk=aid)
                        ArtistGroup.objects.create(group=group, artist=artist_obj, artist_role=(role or ''))
                    except Artist.DoesNotExist:
                        continue
                
                messages.success(request, f'Группа "{name}" успешно обновлена!')
                return redirect('music:admin_groups')
            except Exception as e:
                messages.error(request, f'Ошибка обновления группы: {str(e)}')
        else:
            messages.error(request, 'Заполните все обязательные поля!')
    
    # Данные для формы участников
    artists = Artist.objects.all().order_by('name')
    artist_links = ArtistGroup.objects.filter(group=group).select_related('artist').order_by('artist__name')

    context = {
        'group': group,
        'artists': artists,
        'artist_links': artist_links,
    }
    return render(request, 'music/admin/admin_edit_group.html', context)


@login_required
def admin_delete_group(request, group_id):
    """Удаление группы"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    try:
        group = Group.objects.get(id=group_id)
        group_name = group.name
        group.delete()
        messages.success(request, f'Группа "{group_name}" успешно удалена!')
    except Group.DoesNotExist:
        messages.error(request, 'Группа не найдена!')
    except Exception as e:
        messages.error(request, f'Ошибка удаления группы: {str(e)}')
    
    return redirect('music:admin_groups')


@login_required
def admin_create_group(request):
    """Создание новой группы"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    if request.method == 'POST':
        # Обработка создания группы
        name = request.POST.get('name')
        description = request.POST.get('description')
        photo = request.FILES.get('photo')
        
        if name:
            try:
                group = Group.objects.create(
                    name=name,
                    description=description if description else '',
                    photo=photo if photo else None
                )
                
                messages.success(request, f'Группа "{name}" успешно создана!')
                return redirect('music:admin_groups')
            except Exception as e:
                messages.error(request, f'Ошибка создания группы: {str(e)}')
        else:
            messages.error(request, 'Заполните все обязательные поля!')
    
    context = {}
    return render(request, 'music/admin/admin_create_group.html', context)


@login_required
def admin_genres(request):
    """Управление жанрами"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    genres = Genre.objects.all().order_by('name')
    
    context = {
        'genres': genres,
    }
    return render(request, 'music/admin/admin_genres.html', context)


@login_required
def admin_create_genre(request):
    """Создание жанра"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        
        if name:
            genre = Genre.objects.create(
                name=name
            )
            messages.success(request, f'Жанр "{genre.name}" успешно создан!')
            return redirect('music:admin_genres')
        else:
            messages.error(request, 'Название жанра обязательно!')
    
    return render(request, 'music/admin/admin_create_genre.html')
            

@login_required
def admin_edit_genre(request, pk):
    """Редактирование жанра"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    genre = get_object_or_404(Genre, pk=pk)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        
        if name:
            genre.name = name
            genre.save()
            messages.success(request, f'Жанр "{genre.name}" успешно обновлен!')
            return redirect('music:admin_genres')
        else:
            messages.error(request, 'Название жанра обязательно!')
    
    context = {
        'genre': genre,
    }
    return render(request, 'music/admin/admin_edit_genre.html', context)


@login_required
def admin_delete_genre(request, pk):
    """Удаление жанра"""
    if not request.user.is_authenticated or request.user.role != 'admin':
        messages.error(request, 'Доступ запрещен. Требуются права администратора.')
        return redirect('music:home')
    
    genre = get_object_or_404(Genre, pk=pk)
    
    if request.method == 'POST':
        genre_name = genre.name
        genre.delete()
        messages.success(request, f'Жанр "{genre_name}" успешно удален!')
        return redirect('music:admin_genres')
    
    context = {
        'genre': genre,
    }
    return render(request, 'music/admin/admin_delete_genre.html', context)
